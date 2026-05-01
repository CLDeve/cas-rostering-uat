from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from roster_system.schemas import Gender, ShiftPattern


@dataclass
class ImportedEmployeeRow:
    serial_number: int | None
    team: str
    rank: str
    staff_id: str
    name: str
    start_date: date | None
    gender: Gender
    cert: str | None
    scheme: str
    shift_pattern: ShiftPattern
    contractual_hours: Decimal
    forecast_hours: Decimal


class EmployeeImportError(Exception):
    pass


_HEADER_ALIASES: dict[str, set[str]] = {
    "serial_number": {"sn", "sno", "serialnumber", "serialno"},
    "team": {"team"},
    "rank": {"rank"},
    "staff_id": {"id", "staffid", "employeeid"},
    "name": {"name"},
    "start_date": {"startdate", "officerstartdate"},
    "gender": {"gender", "sex"},
    "cert": {"cert"},
    "scheme": {"scheme"},
    "shift_pattern": {"shiftpattern", "shift"},
    "contractual_hours": {"contractualhours", "contracthours"},
    "forecast_hours": {"forecasthours"},
}

_REQUIRED_TEMPLATE_FIELDS = {
    "team",
    "rank",
    "staff_id",
    "name",
    "start_date",
    "gender",
    "scheme",
    "contractual_hours",
}


def _require_non_blank(value: object, field_name: str, row_num: int) -> str:
    cleaned = _as_clean_str(value)
    if not cleaned:
        raise EmployeeImportError(f"Missing required value for {field_name} at row {row_num}")
    return cleaned


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]", "", text)


def _to_decimal(value: object, field_name: str) -> Decimal:
    if value is None or value == "":
        raise EmployeeImportError(f"Missing required value for {field_name}")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError) as exc:
        raise EmployeeImportError(f"Invalid decimal value for {field_name}: {value}") from exc


def _to_int(value: object, field_name: str) -> int:
    if value is None or value == "":
        raise EmployeeImportError(f"Missing required value for {field_name}")
    try:
        return int(value)
    except (ValueError, TypeError) as exc:
        raise EmployeeImportError(f"Invalid integer value for {field_name}: {value}") from exc


def _as_clean_str(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _to_date(value: object, field_name: str, row_num: int) -> date:
    if value is None or value == "":
        raise EmployeeImportError(f"Missing required value for {field_name} at row {row_num}")

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _as_clean_str(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise EmployeeImportError(
        f"Invalid date format for {field_name} at row {row_num}. Use YYYY-MM-DD."
    )


def _infer_shift_pattern(scheme: str) -> ShiftPattern:
    return ShiftPattern.WORK_4_OFF_2 if scheme.upper() == "B" else ShiftPattern.WORK_5_OFF_1


def _parse_shift_pattern(value: object, scheme: str) -> ShiftPattern:
    raw = _as_clean_str(value)
    if not raw:
        return _infer_shift_pattern(scheme)

    upper = raw.upper()
    if upper == ShiftPattern.WORK_4_OFF_2.value:
        return ShiftPattern.WORK_4_OFF_2
    if upper == ShiftPattern.WORK_5_OFF_1.value:
        return ShiftPattern.WORK_5_OFF_1
    raise EmployeeImportError(f"Invalid shift_pattern: {raw}. Allowed: 4W2O, 5W1O")


def _parse_gender(value: object) -> Gender:
    raw = _as_clean_str(value)
    if not raw:
        return Gender.UNKNOWN
    upper = raw.upper()
    if upper in {"M", "MALE"}:
        return Gender.MALE
    if upper in {"F", "FEMALE"}:
        return Gender.FEMALE
    if upper in {"OTHER", "O"}:
        return Gender.OTHER
    if upper == "UNKNOWN":
        return Gender.UNKNOWN
    raise EmployeeImportError(f"Invalid gender: {raw}. Allowed: MALE/FEMALE/OTHER/UNKNOWN")


def _extract_rows_from_sap_layout(sheet: Worksheet) -> list[ImportedEmployeeRow]:
    imported: list[ImportedEmployeeRow] = []
    last_team = ""

    for row_num in range(4, sheet.max_row + 1):
        serial_raw = sheet[f"A{row_num}"].value
        staff_id_raw = sheet[f"D{row_num}"].value
        name_raw = sheet[f"E{row_num}"].value

        if serial_raw is None and not _as_clean_str(name_raw):
            continue

        team_raw = _as_clean_str(sheet[f"B{row_num}"].value)
        if team_raw:
            last_team = team_raw

        rank = _as_clean_str(sheet[f"C{row_num}"].value)
        staff_id = _as_clean_str(staff_id_raw)
        name = _as_clean_str(name_raw)

        if name in {"LEGEND"}:
            break
        if staff_id == "" and rank == "" and name == "":
            continue

        if not staff_id or not name or not rank or not last_team:
            continue

        cert = _as_clean_str(sheet[f"F{row_num}"].value) or None
        scheme = _as_clean_str(sheet[f"AQ{row_num}"].value) or "A"
        shift_pattern = _infer_shift_pattern(scheme)

        contractual_hours = _to_decimal(sheet[f"AO{row_num}"].value, "contractual_hours")
        overtime_hours = _to_decimal(sheet[f"AP{row_num}"].value, "ot")
        forecast_hours = (contractual_hours + overtime_hours).quantize(Decimal("0.01"))
        if forecast_hours < 0:
            forecast_hours = Decimal("0.00")

        imported.append(
            ImportedEmployeeRow(
                serial_number=_to_int(serial_raw, "serial_number") if serial_raw not in (None, "") else None,
                team=last_team,
                rank=rank,
                staff_id=staff_id,
                name=name,
                start_date=None,
                gender=Gender.UNKNOWN,
                cert=cert,
                scheme=scheme,
                shift_pattern=shift_pattern,
                contractual_hours=contractual_hours,
                forecast_hours=forecast_hours,
            )
        )

    if not imported:
        raise EmployeeImportError("No employee rows found in selected sheet")

    return imported


def _header_mapping(sheet: Worksheet) -> dict[str, int] | None:
    headers = [sheet.cell(row=1, column=i).value for i in range(1, min(sheet.max_column, 30) + 1)]
    normalized = [_normalize_header(h) for h in headers]

    mapping: dict[str, int] = {}
    for idx, key in enumerate(normalized, start=1):
        if not key:
            continue
        for canonical, aliases in _HEADER_ALIASES.items():
            if key in aliases and canonical not in mapping:
                mapping[canonical] = idx
                break

    if _REQUIRED_TEMPLATE_FIELDS.issubset(mapping.keys()):
        return mapping
    return None


def _extract_rows_from_template_layout(sheet: Worksheet, mapping: dict[str, int]) -> list[ImportedEmployeeRow]:
    imported: list[ImportedEmployeeRow] = []

    for row_num in range(2, sheet.max_row + 1):
        required_values = [
            sheet.cell(row=row_num, column=mapping[field]).value
            for field in _REQUIRED_TEMPLATE_FIELDS
        ]
        if all((v is None or _as_clean_str(v) == "") for v in required_values):
            continue

        team = _require_non_blank(sheet.cell(row=row_num, column=mapping["team"]).value, "team", row_num)
        rank = _require_non_blank(sheet.cell(row=row_num, column=mapping["rank"]).value, "rank", row_num)
        staff_id = _require_non_blank(sheet.cell(row=row_num, column=mapping["staff_id"]).value, "staff_id", row_num)
        name = _require_non_blank(sheet.cell(row=row_num, column=mapping["name"]).value, "name", row_num)
        scheme = _require_non_blank(sheet.cell(row=row_num, column=mapping["scheme"]).value, "scheme", row_num)

        shift_raw = sheet.cell(row=row_num, column=mapping["shift_pattern"]).value if "shift_pattern" in mapping else None
        shift_pattern = _parse_shift_pattern(shift_raw, scheme)

        cert_value = None
        if "cert" in mapping:
            cert_value = _as_clean_str(sheet.cell(row=row_num, column=mapping["cert"]).value) or None

        imported.append(
            ImportedEmployeeRow(
                serial_number=(
                    _to_int(sheet.cell(row=row_num, column=mapping["serial_number"]).value, "serial_number")
                    if "serial_number" in mapping and sheet.cell(row=row_num, column=mapping["serial_number"]).value not in (None, "")
                    else None
                ),
                team=team,
                rank=rank,
                staff_id=staff_id,
                name=name,
                start_date=_to_date(sheet.cell(row=row_num, column=mapping["start_date"]).value, "start_date", row_num),
                gender=_parse_gender(sheet.cell(row=row_num, column=mapping["gender"]).value),
                cert=cert_value,
                scheme=scheme,
                shift_pattern=shift_pattern,
                contractual_hours=_to_decimal(sheet.cell(row=row_num, column=mapping["contractual_hours"]).value, "contractual_hours"),
                forecast_hours=(
                    _to_decimal(sheet.cell(row=row_num, column=mapping["forecast_hours"]).value, "forecast_hours")
                    if "forecast_hours" in mapping
                    else _to_decimal(sheet.cell(row=row_num, column=mapping["contractual_hours"]).value, "contractual_hours")
                ),
            )
        )

    if not imported:
        raise EmployeeImportError("No employee rows found in selected sheet")

    return imported


def parse_employee_rows_from_excel(file_bytes: bytes) -> tuple[list[ImportedEmployeeRow], str]:
    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except (OSError, ValueError, TypeError, BadZipFile) as exc:
        raise EmployeeImportError("Unable to read Excel file") from exc

    parse_errors: list[str] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        mapping = _header_mapping(sheet)
        try:
            rows = (
                _extract_rows_from_template_layout(sheet, mapping)
                if mapping is not None
                else _extract_rows_from_sap_layout(sheet)
            )
            return rows, sheet_name
        except EmployeeImportError as exc:
            parse_errors.append(f"{sheet_name}: {exc}")
            continue

    if parse_errors:
        raise EmployeeImportError("Unable to parse any worksheet. " + " | ".join(parse_errors))
    raise EmployeeImportError("Excel file has no worksheets")
