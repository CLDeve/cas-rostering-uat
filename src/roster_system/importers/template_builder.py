from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

_TEMPLATE_HEADERS = [
    "TEAM",
    "RANK",
    "ID",
    "NAME",
    "Start Date",
    "Gender",
    "CERT",
    "Scheme",
    "Shift Pattern",
    "Contractual hours",
]

_EXAMPLE_ROW = [
    "A1",
    "SGT",
    "100001",
    "JOHN DOE",
    "2026-02-01",
    "MALE",
    "ADP",
    "A",
    "5W1O",
    264,
]


def build_employee_upload_template(sheet_name: str = "SAP FEB (AM)") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    for col_index, header in enumerate(_TEMPLATE_HEADERS, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_index, value in enumerate(_EXAMPLE_ROW, start=1):
        sheet.cell(row=2, column=col_index, value=value)

    sheet.freeze_panes = "A2"
    column_widths = [10, 10, 12, 28, 14, 12, 12, 10, 14, 18]
    for idx, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(ord('A') + idx - 1)].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
