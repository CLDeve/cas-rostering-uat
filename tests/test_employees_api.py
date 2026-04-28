from io import BytesIO
from decimal import Decimal

from openpyxl import Workbook, load_workbook
from fastapi.testclient import TestClient

from roster_system.main import app


def employee_payload(staff_id: str = "148928") -> dict:
    return {
        "team": "A1",
        "rank": "SGT2",
        "staff_id": staff_id,
        "name": "SAMUEL TEE HONG",
        "start_date": "2026-02-01",
        "gender": "MALE",
        "cert": "ADP",
        "scheme": "A",
        "shift_pattern": "5W1O",
        "contractual_hours": "264.00",
        "forecast_hours": "287.00",
    }


def make_upload_file() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "SAP FEB (AM)"

    # Minimal row following the expected import shape.
    ws["A4"] = 1
    ws["B4"] = "A1"
    ws["C4"] = "SGT2"
    ws["D4"] = "148928"
    ws["E4"] = "SAMUEL TEE HONG"
    ws["F4"] = "ADP"
    ws["AO4"] = 264
    ws["AP4"] = 40
    ws["AQ4"] = "A"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_create_employee() -> None:
    with TestClient(app) as client:
        response = client.post("/api/v1/employees", json=employee_payload("100001"))

    assert response.status_code == 201
    body = response.json()
    assert body["staff_id"] == "100001"
    assert body["serial_number"] == 1
    assert Decimal(body["contractual_hours"]) == Decimal("264.00")


def test_duplicate_staff_id_conflict() -> None:
    with TestClient(app) as client:
        first = client.post("/api/v1/employees", json=employee_payload("100002"))
        second = client.post("/api/v1/employees", json=employee_payload("100002"))

    assert first.status_code == 201
    assert second.status_code == 409


def test_list_employees_with_filter() -> None:
    with TestClient(app) as client:
        client.post("/api/v1/employees", json=employee_payload("100003"))
        client.post(
            "/api/v1/employees",
            json={
                **employee_payload("100004"),
                "team": "A2",
                "scheme": "B",
            },
        )
        response = client.get("/api/v1/employees", params={"team": "A2", "scheme": "B"})

    assert response.status_code == 200
    body = response.json()
    assert body["total"] >= 1
    assert all(item["team"] == "A2" and item["scheme"] == "B" for item in body["items"])


def test_update_employee() -> None:
    with TestClient(app) as client:
        created = client.post("/api/v1/employees", json=employee_payload("100005")).json()
        response = client.put(
            f"/api/v1/employees/{created['id']}",
            json={"forecast_hours": "300.00"},
        )

    assert response.status_code == 200
    assert Decimal(response.json()["forecast_hours"]) == Decimal("300.00")


def test_delete_employee() -> None:
    with TestClient(app) as client:
        created = client.post("/api/v1/employees", json=employee_payload("100006")).json()
        delete_response = client.delete(f"/api/v1/employees/{created['id']}")
        get_response = client.get(f"/api/v1/employees/{created['id']}")

    assert delete_response.status_code == 204
    assert get_response.status_code == 404


def test_get_shift_plan() -> None:
    with TestClient(app) as client:
        created = client.post("/api/v1/employees", json=employee_payload("100007")).json()
        response = client.get(f"/api/v1/employees/{created['id']}/shift-plan", params={"days": 6})

    assert response.status_code == 200
    payload = response.json()
    assert payload["shift_pattern"] == "5W1O"
    assert payload["work_days"] == 5
    assert payload["off_days"] == 1


def test_download_upload_template() -> None:
    with TestClient(app) as client:
        response = client.get("/api/v1/employees/upload-template", params={"sheet_name": "SAP FEB (AM)"})

    assert response.status_code == 200
    assert "attachment; filename=\"employee_upload_template.xlsx\"" in response.headers.get("content-disposition", "")


def test_upload_employees_from_template() -> None:
    with TestClient(app) as client:
        template_response = client.get("/api/v1/employees/upload-template", params={"sheet_name": "SAP FEB (AM)"})
        assert template_response.status_code == 200

        workbook = load_workbook(filename=BytesIO(template_response.content))
        sheet = workbook["SAP FEB (AM)"]
        sheet["A2"] = "A1"
        sheet["B2"] = "SGT"
        sheet["C2"] = "200001"
        sheet["D2"] = "TEMPLATE USER"
        sheet["E2"] = "2026-02-01"
        sheet["F2"] = "FEMALE"
        sheet["G2"] = "ADP"
        sheet["H2"] = "A"
        sheet["I2"] = "5W1O"
        sheet["J2"] = 264
        buffer = BytesIO()
        workbook.save(buffer)

        response = client.post(
            "/api/v1/employees/upload",
            files={"file": ("template_filled.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"sheet_name": "SAP FEB (AM)"},
        )

        list_response = client.get("/api/v1/employees")

    assert response.status_code == 200
    assert list_response.status_code == 200
    payload = list_response.json()
    assert payload["total"] == 1
    assert payload["items"][0]["staff_id"] == "200001"
    assert payload["items"][0]["serial_number"] == 1
    assert payload["items"][0]["gender"] == "FEMALE"


def test_upload_template_missing_required_field_rejected() -> None:
    with TestClient(app) as client:
        template_response = client.get("/api/v1/employees/upload-template", params={"sheet_name": "SAP FEB (AM)"})
        assert template_response.status_code == 200

        workbook = load_workbook(filename=BytesIO(template_response.content))
        sheet = workbook["SAP FEB (AM)"]
        sheet["A2"] = "A1"
        sheet["B2"] = "SGT"
        sheet["C2"] = ""  # Missing ID
        sheet["D2"] = "TEMPLATE USER"
        sheet["E2"] = "2026-02-01"
        sheet["F2"] = "FEMALE"
        sheet["H2"] = "A"
        sheet["J2"] = 264
        buffer = BytesIO()
        workbook.save(buffer)

        response = client.post(
            "/api/v1/employees/upload",
            files={"file": ("template_missing_id.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"sheet_name": "SAP FEB (AM)"},
        )

    assert response.status_code == 400
    assert "staff_id" in response.json()["detail"]


def test_upload_template_invalid_gender_rejected() -> None:
    with TestClient(app) as client:
        template_response = client.get("/api/v1/employees/upload-template", params={"sheet_name": "SAP FEB (AM)"})
        assert template_response.status_code == 200

        workbook = load_workbook(filename=BytesIO(template_response.content))
        sheet = workbook["SAP FEB (AM)"]
        sheet["A2"] = "A1"
        sheet["B2"] = "SGT"
        sheet["C2"] = "200099"
        sheet["D2"] = "TEMPLATE USER"
        sheet["E2"] = "2026-02-01"
        sheet["F2"] = "INVALID"
        sheet["H2"] = "A"
        sheet["J2"] = 264
        buffer = BytesIO()
        workbook.save(buffer)

        response = client.post(
            "/api/v1/employees/upload",
            files={"file": ("template_invalid_gender.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"sheet_name": "SAP FEB (AM)"},
        )

    assert response.status_code == 400
    assert "Invalid gender" in response.json()["detail"]


def test_upload_employees() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/v1/employees/upload",
            files={"file": ("roster.xlsx", make_upload_file(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"sheet_name": "SAP FEB (AM)"},
        )

        list_response = client.get("/api/v1/employees")

    assert response.status_code == 200
    payload = response.json()
    assert payload["created"] == 1
    assert payload["updated"] == 0
    assert payload["total_processed"] == 1
    assert payload["upload_file_id"] is not None
    assert payload["download_url"] is not None

    assert list_response.status_code == 200
    assert list_response.json()["total"] == 1


def test_download_latest_uploaded_file() -> None:
    with TestClient(app) as client:
        upload_response = client.post(
            "/api/v1/employees/upload",
            files={"file": ("roster.xlsx", make_upload_file(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"sheet_name": "SAP FEB (AM)"},
        )
        assert upload_response.status_code == 200

        metadata_response = client.get("/api/v1/employees/upload-files/latest")
        download_response = client.get("/api/v1/employees/upload-files/latest/download")

    assert metadata_response.status_code == 200
    assert metadata_response.json()["original_filename"] == "roster.xlsx"
    assert download_response.status_code == 200
    assert "attachment; filename=\"roster.xlsx\"" in download_response.headers.get("content-disposition", "")


def test_roster_calendar_generation() -> None:
    with TestClient(app) as client:
        create_response = client.post(
            "/api/v1/employees",
            json={
                **employee_payload("300001"),
                "shift_pattern": "4W2O",
            },
        )
        assert create_response.status_code == 201

        calendar_response = client.get("/api/v1/roster/calendar", params={"year": 2026, "month": 2})

    assert calendar_response.status_code == 200
    payload = calendar_response.json()
    assert payload["year"] == 2026
    assert payload["month"] == 2
    assert payload["days_in_month"] == 28
    assert len(payload["day_headers"]) == 28
    assert len(payload["employees"]) == 1
    assert len(payload["employees"][0]["schedule"]) == 28
    assert payload["employees"][0]["forecast_hours"] == 260


def test_roster_calendar_pre_start_date_cells_empty() -> None:
    with TestClient(app) as client:
        create_response = client.post(
            "/api/v1/employees",
            json={
                **employee_payload("300002"),
                "start_date": "2026-02-10",
                "shift_pattern": "5W1O",
            },
        )
        assert create_response.status_code == 201

        calendar_response = client.get("/api/v1/roster/calendar", params={"year": 2026, "month": 2})

    assert calendar_response.status_code == 200
    payload = calendar_response.json()
    schedule = payload["employees"][0]["schedule"]
    assert schedule[0] == "EMPTY"
    assert schedule[8] == "EMPTY"
    assert schedule[9] in {"WORK", "OFF"}


def test_roster_calendar_save_persists_overrides() -> None:
    with TestClient(app) as client:
        create_response = client.post(
            "/api/v1/employees",
            json={
                **employee_payload("300003"),
                "shift_pattern": "5W1O",
            },
        )
        assert create_response.status_code == 201

        first_calendar = client.get("/api/v1/roster/calendar", params={"year": 2026, "month": 2})
        assert first_calendar.status_code == 200
        payload = first_calendar.json()
        employee_row = payload["employees"][0]
        schedule = list(employee_row["schedule"])
        schedule[0] = "OT1"
        schedule[1] = "OT2"

        save_response = client.put(
            "/api/v1/roster/calendar",
            json={
                "year": 2026,
                "month": 2,
                "employees": [
                    {
                        "employee_id": employee_row["employee_id"],
                        "schedule": schedule,
                    }
                ],
            },
        )
        second_calendar = client.get("/api/v1/roster/calendar", params={"year": 2026, "month": 2})

    assert save_response.status_code == 200
    saved_payload = save_response.json()
    assert saved_payload["employees_saved"] == 1
    assert saved_payload["overrides_saved"] >= 1

    assert second_calendar.status_code == 200
    updated_schedule = second_calendar.json()["employees"][0]["schedule"]
    assert updated_schedule[0] == "OT1"
    assert updated_schedule[1] == "OT2"


def test_roster_calendar_save_blocks_more_than_12_consecutive_working_days() -> None:
    with TestClient(app) as client:
        create_response = client.post(
            "/api/v1/employees",
            json={
                **employee_payload("300004"),
                "shift_pattern": "5W1O",
            },
        )
        assert create_response.status_code == 201

        calendar_response = client.get("/api/v1/roster/calendar", params={"year": 2026, "month": 2})
        assert calendar_response.status_code == 200
        payload = calendar_response.json()
        employee_row = payload["employees"][0]

        forced_working = ["WORK"] * payload["days_in_month"]
        save_response = client.put(
            "/api/v1/roster/calendar",
            json={
                "year": 2026,
                "month": 2,
                "employees": [
                    {
                        "employee_id": employee_row["employee_id"],
                        "schedule": forced_working,
                    }
                ],
            },
        )

    assert save_response.status_code == 422
    assert "12 consecutive working days" in save_response.json()["detail"]


def test_create_and_list_deployment_site() -> None:
    with TestClient(app) as client:
        create_response = client.post(
            "/api/v1/deployments",
            json={
                "site_name": "Terminal 3 Landside",
                "deployment_days": ["MON", "TUE", "WED", "THU", "FRI"],
                "requirements": [
                    {
                        "product_type": "APO",
                        "required_headcount": 12,
                        "reporting_from": "08:00",
                        "reporting_to": "16:00",
                        "next_shift_from": "16:00",
                        "next_shift_to": "00:00",
                    },
                    {
                        "product_type": "AVSO",
                        "required_headcount": 4,
                        "reporting_from": "07:00",
                        "reporting_to": "15:00",
                        "next_shift_from": "15:00",
                        "next_shift_to": "23:00",
                    },
                ],
            },
        )
        list_response = client.get("/api/v1/deployments")

    assert create_response.status_code == 201
    assert list_response.status_code == 200
    payload = list_response.json()
    assert len(payload) == 1
    assert payload[0]["site_name"] == "Terminal 3 Landside"
    assert payload[0]["deployment_days"] == ["MON", "TUE", "WED", "THU", "FRI"]
    assert len(payload[0]["requirements"]) == 2
    assert payload[0]["requirements"][0]["product_type"] == "APO"
    assert payload[0]["requirements"][1]["product_type"] == "AVSO"


def test_create_adhoc_deployment_site() -> None:
    with TestClient(app) as client:
        create_response = client.post(
            "/api/v1/deployments",
            json={
                "site_name": "Event Site Alpha",
                "mode": "ADHOC",
                "deployment_days": [],
                "adhoc_start_at": "2026-05-01T08:00:00",
                "adhoc_end_at": "2026-05-01T20:00:00",
                "requirements": [
                    {
                        "product_type": "APO",
                        "required_headcount": 6,
                        "reporting_from": "08:00",
                        "reporting_to": "14:00",
                        "next_shift_from": "14:00",
                        "next_shift_to": "20:00",
                    }
                ],
            },
        )
        list_response = client.get("/api/v1/deployments")

    assert create_response.status_code == 201
    assert list_response.status_code == 200
    payload = list_response.json()
    assert len(payload) == 1
    assert payload[0]["mode"] == "ADHOC"
    assert payload[0]["deployment_days"] == []


def test_replace_and_list_deployment_assignments() -> None:
    with TestClient(app) as client:
        site_resp = client.post(
            "/api/v1/deployments",
            json={
                "site_name": "Terminal 2 Screening",
                "deployment_days": ["MON", "TUE"],
                "requirements": [
                    {
                        "product_type": "APO",
                        "required_headcount": 2,
                        "reporting_from": "08:00",
                        "reporting_to": "16:00",
                        "next_shift_from": "16:00",
                        "next_shift_to": "00:00",
                    }
                ],
            },
        )
        assert site_resp.status_code == 201
        site_id = site_resp.json()["id"]

        emp1 = client.post("/api/v1/employees", json=employee_payload("900001"))
        emp2 = client.post("/api/v1/employees", json=employee_payload("900002"))
        assert emp1.status_code == 201
        assert emp2.status_code == 201

        replace_response = client.put(
            "/api/v1/deployments/assignments",
            json={
                "deployment_date": "2026-05-02",
                "assignments": [
                    {"site_id": site_id, "slot_index": 0, "employee_id": emp1.json()["id"]},
                    {"site_id": site_id, "slot_index": 1, "employee_id": emp2.json()["id"]},
                ],
            },
        )
        get_response = client.get(
            "/api/v1/deployments/assignments",
            params={"deployment_date": "2026-05-02"},
        )

    assert replace_response.status_code == 200
    assert get_response.status_code == 200
    payload = get_response.json()
    assert payload["deployment_date"] == "2026-05-02"
    assert len(payload["assignments"]) == 2
    assert payload["assignments"][0]["site_id"] == site_id
    assert payload["assignments"][0]["slot_index"] == 0


def test_replace_deployment_assignments_invalid_employee_rejected() -> None:
    with TestClient(app) as client:
        site_resp = client.post(
            "/api/v1/deployments",
            json={
                "site_name": "Terminal 1 Entry",
                "deployment_days": ["MON"],
                "requirements": [
                    {
                        "product_type": "AVSO",
                        "required_headcount": 1,
                        "reporting_from": "08:00",
                        "reporting_to": "16:00",
                        "next_shift_from": "16:00",
                        "next_shift_to": "00:00",
                    }
                ],
            },
        )
        assert site_resp.status_code == 201
        site_id = site_resp.json()["id"]

        replace_response = client.put(
            "/api/v1/deployments/assignments",
            json={
                "deployment_date": "2026-05-03",
                "assignments": [
                    {"site_id": site_id, "slot_index": 0, "employee_id": 999999},
                ],
            },
        )

    assert replace_response.status_code == 422
    assert "employee_id" in replace_response.json()["detail"]


def test_dashboard_coverage_daily_and_calendar() -> None:
    with TestClient(app) as client:
        site_resp = client.post(
            "/api/v1/deployments",
            json={
                "site_name": "Terminal 4 Gates",
                "mode": "RECURRING",
                "deployment_days": ["MON"],
                "requirements": [
                    {
                        "product_type": "APO",
                        "required_headcount": 2,
                        "reporting_from": "08:00",
                        "reporting_to": "16:00",
                        "next_shift_from": "16:00",
                        "next_shift_to": "00:00",
                    }
                ],
            },
        )
        assert site_resp.status_code == 201
        site_id = site_resp.json()["id"]

        emp1 = client.post("/api/v1/employees", json=employee_payload("920001"))
        emp2 = client.post("/api/v1/employees", json=employee_payload("920002"))
        assert emp1.status_code == 201
        assert emp2.status_code == 201

        assign_resp = client.put(
            "/api/v1/deployments/assignments",
            json={
                "deployment_date": "2026-04-27",
                "assignments": [
                    {"site_id": site_id, "slot_index": 0, "employee_id": emp1.json()["id"]},
                    {"site_id": site_id, "slot_index": 1, "employee_id": emp2.json()["id"]},
                ],
            },
        )
        assert assign_resp.status_code == 200

        daily_resp = client.get("/api/v1/dashboard/coverage", params={"date": "2026-04-27"})
        calendar_resp = client.get("/api/v1/dashboard/coverage-calendar", params={"year": 2026, "month": 4})

    assert daily_resp.status_code == 200
    daily_payload = daily_resp.json()
    assert daily_payload["date"] == "2026-04-27"
    assert daily_payload["required_headcount"] == 2
    assert daily_payload["assigned_headcount"] == 2
    assert daily_payload["is_covered"] is True

    assert calendar_resp.status_code == 200
    calendar_payload = calendar_resp.json()
    assert calendar_payload["year"] == 2026
    assert calendar_payload["month"] == 4
    assert len(calendar_payload["days"]) == 30
    apr27 = [row for row in calendar_payload["days"] if row["date"] == "2026-04-27"][0]
    assert apr27["required_headcount"] == 2
    assert apr27["assigned_headcount"] == 2


def test_create_list_and_update_user_management() -> None:
    with TestClient(app) as client:
        create_response = client.post(
            "/api/v1/users",
            json={
                "staff_id": "U1001",
                "username": "charles.liew",
                "display_name": "Charles Liew",
                "role": "PLANNER",
            },
        )
        list_response = client.get("/api/v1/users")

        assert create_response.status_code == 201
        created_user = create_response.json()
        assert created_user["is_active"] is True

        update_response = client.put(
            f"/api/v1/users/{created_user['id']}/status",
            json={"is_active": False},
        )
        list_after_response = client.get("/api/v1/users")

    assert list_response.status_code == 200
    assert len(list_response.json()) == 1
    assert list_response.json()[0]["username"] == "charles.liew"

    assert update_response.status_code == 200
    assert update_response.json()["is_active"] is False
    assert list_after_response.status_code == 200
    assert list_after_response.json()[0]["is_active"] is False


def test_user_management_duplicate_username_conflict() -> None:
    with TestClient(app) as client:
        first = client.post(
            "/api/v1/users",
            json={
                "staff_id": "U2001",
                "username": "shared.username",
                "display_name": "User One",
                "role": "VIEWER",
            },
        )
        second = client.post(
            "/api/v1/users",
            json={
                "staff_id": "U2002",
                "username": "shared.username",
                "display_name": "User Two",
                "role": "ADMIN",
            },
        )

    assert first.status_code == 201
    assert second.status_code == 409


def test_user_management_duplicate_staff_id_conflict() -> None:
    with TestClient(app) as client:
        first = client.post(
            "/api/v1/users",
            json={
                "staff_id": "U3001",
                "username": "alpha.user",
                "display_name": "Alpha User",
                "role": "VIEWER",
            },
        )
        second = client.post(
            "/api/v1/users",
            json={
                "staff_id": "U3001",
                "username": "beta.user",
                "display_name": "Beta User",
                "role": "PLANNER",
            },
        )

    assert first.status_code == 201
    assert second.status_code == 409


def test_create_and_list_training_course() -> None:
    with TestClient(app) as client:
        create_response = client.post(
            "/api/v1/trainings",
            json={
                "course_name": "Aviation Security Refresher",
                "location": "Training Room T3-2A",
                "start_at": "2026-05-20T09:00:00+08:00",
                "end_at": "2026-05-20T17:00:00+08:00",
            },
        )
        list_response = client.get("/api/v1/trainings")

    assert create_response.status_code == 201
    assert list_response.status_code == 200
    payload = list_response.json()
    assert len(payload) == 1
    assert payload[0]["course_name"] == "Aviation Security Refresher"
    assert payload[0]["location"] == "Training Room T3-2A"


def test_create_training_course_invalid_window_rejected() -> None:
    with TestClient(app) as client:
        create_response = client.post(
            "/api/v1/trainings",
            json={
                "course_name": "Incident Command Drill",
                "location": "Ops Classroom",
                "start_at": "2026-05-21T14:00:00+08:00",
                "end_at": "2026-05-21T10:00:00+08:00",
            },
        )

    assert create_response.status_code == 422
    assert "end_at must be later than start_at" in str(create_response.json())
